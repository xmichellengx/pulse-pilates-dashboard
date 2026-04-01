#!/usr/bin/env python3
"""
Import historical order data from Pilates Machine Order Tracker.xlsx into Supabase.
Reads sheets: Customer order 2024, Customer order 2025, Customer order 2026
"""

import json
import re
import sys
import datetime
import requests
import openpyxl

XLSX_PATH = "/Users/michelleng/Downloads/Pilates Machine Order Tracker.xlsx"
SUPABASE_URL = "https://ykexbcyrhknswxaabjyn.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InlrZXhiY3lyaGtuc3d4YWFianluIiwicm9sZSI6"
    "InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3Mjg3MTUwMywiZXhwIjoyMDg4NDQ3NTAzfQ."
    "LNpGDd__pxuvPV3x_Yj9DRSgLLn3K03FtgN0OBlhFlM"
)
BATCH_SIZE = 50

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates",
}

SHEET_NAMES = ["Customer order 2024", "Customer order 2025", "Customer order 2026"]

# Column indices (0-based) based on header row
COL = {
    "name": 0,
    "email": 1,
    "product": 2,
    "leads": 3,
    "unit": 4,
    "mode": 5,
    "duration": 6,
    "payment": 7,
    "amount": 8,
    "monthly_rental": 9,
    "balance": 10,
    "payment_date": 11,
    "phone": 12,
    "payex_status": 13,
    "delivery_date": 14,
    "location": 15,
    "address": 16,
    "status": 17,
    "invoice": 18,
    "case_code": 19,
    "bill_booking": 20,
    "bill_delivery": 21,
    "remarks": 22,
}


def normalise_status(raw: str) -> str:
    if not raw:
        return "Pending"
    s = str(raw).strip()
    mapping = {
        "Delivered": "Delivered",
        "Delivered ": "Delivered",
        "Cancel Order": "Cancelled",
        "Cancelled": "Cancelled",
        "Pending Delivered": "Pending Delivered",
        "Returned": "Returned",
        "Returned ": "Returned",
    }
    return mapping.get(s, s)


def normalise_lead_source(raw: str) -> str:
    if not raw:
        return None
    s = str(raw).strip()
    mapping = {
        "Google": "Google",
        "Facebook": "Facebook",
        "Instagram": "Instagram",
        "Tiktok": "TikTok",
        "TikTok": "TikTok",
        "XHS": "XHS",
        "Referral": "Referral",
        "Michelle Referral": "Referral",
        "Referral (Aisy's customer)": "Referral",
        "SGS Carousel (Aisy)": "Referral",
        "Carousel (Aisy)": "Referral",
        "SGS Facebook (Aisy)": "Facebook",
        "Repeat Customer": "Repeat Customer",
        "Walk In": "Walk In",
        "Shopee": "Shopee",
        "Shopify": "Shopify",
    }
    return mapping.get(s, s)


def normalise_mode(raw: str) -> str:
    """Collapse many mode variants into clean canonical values."""
    if not raw:
        return None
    s = str(raw).strip()
    sl = s.lower()
    if "rental" in sl:
        return "Rental"
    if "shopee" in sl:
        return "Direct Purchase"
    if "returned" in sl:
        return "Direct Purchase"
    if "p4b" in sl:
        return "P4B"
    if "converted" in sl:
        return "Direct Purchase"
    if "purchase" in sl:
        return "Direct Purchase"
    if "cancelled" in sl or "cancel" in sl:
        return "Direct Purchase"
    if "kol" in sl:
        return "KOL"
    return s


def parse_amount(raw):
    """Parse amount — may be a number, a formula string like =4500+150+80, or None."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s or s == "0":
        return 0.0
    # formula like =4500+150+80
    if s.startswith("="):
        expr = s[1:]
        # keep only digits, +, -, *, /, ., (,  )
        safe = re.sub(r"[^0-9+\-*/().]", "", expr)
        try:
            return float(eval(safe))  # noqa: S307
        except Exception:
            return None
    try:
        # Remove commas and parse
        return float(s.replace(",", ""))
    except Exception:
        return None


def parse_date(raw):
    """Return ISO date string or None."""
    if raw is None:
        return None
    if isinstance(raw, (datetime.datetime, datetime.date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_phone(raw):
    if raw is None:
        return None
    if isinstance(raw, float):
        # Excel sometimes stores as float like 60173109651.0
        return str(int(raw))
    s = str(raw).strip()
    # Remove formula prefix
    if s.startswith("="):
        s = s[1:]
    return s if s else None


def determine_market(case_code: str) -> str:
    if not case_code:
        return "MY"
    if "SG" in str(case_code).upper():
        return "SG"
    return "MY"


def get_val(row, key):
    idx = COL[key]
    if idx < len(row):
        return row[idx]
    return None


def row_to_order(row):
    """Convert a spreadsheet row to an orders dict. Returns None if row is empty."""
    name = get_val(row, "name")
    if not name or str(name).strip() == "":
        return None

    case_code_raw = get_val(row, "case_code")
    case_code = str(case_code_raw).strip() if case_code_raw else None

    amount = parse_amount(get_val(row, "amount"))
    monthly = parse_amount(get_val(row, "monthly_rental"))
    balance = parse_amount(get_val(row, "balance"))

    record = {
        "customer_name": str(name).strip(),
        "email": str(get_val(row, "email")).strip() if get_val(row, "email") else None,
        "product_name": str(get_val(row, "product")).strip() if get_val(row, "product") else None,
        "lead_source": normalise_lead_source(get_val(row, "leads")),
        "units": int(get_val(row, "unit")) if get_val(row, "unit") and str(get_val(row, "unit")).replace(".","").isdigit() else 1,
        "mode": normalise_mode(get_val(row, "mode")),
        "payment_type": str(get_val(row, "payment")).strip() if get_val(row, "payment") else None,
        "amount": amount,
        "monthly_rental": monthly if monthly and monthly > 0 else None,
        "balance": balance if balance is not None else 0,
        "payment_date": parse_date(get_val(row, "payment_date")),
        "phone": parse_phone(get_val(row, "phone")),
        "payex_status": str(get_val(row, "payex_status")).strip() if get_val(row, "payex_status") else None,
        "delivery_date": parse_date(get_val(row, "delivery_date")),
        "location": str(get_val(row, "location")).strip() if get_val(row, "location") else None,
        "address": str(get_val(row, "address")).strip() if get_val(row, "address") else None,
        "status": normalise_status(get_val(row, "status")),
        "invoice_sent": bool(get_val(row, "invoice")),
        "case_code": case_code,
        "market": determine_market(case_code),
        "remarks": str(get_val(row, "remarks")).strip() if get_val(row, "remarks") else None,
    }

    # Remove None values so Supabase uses column defaults
    return {k: v for k, v in record.items() if v is not None}


ALL_KEYS = [
    "case_code", "customer_name", "email", "phone", "product_name",
    "units", "mode", "payment_type", "amount", "monthly_rental",
    "balance", "payment_date", "delivery_date", "location", "address",
    "status", "lead_source", "market", "invoice_sent", "payex_status", "remarks",
]


def normalise_keys(record):
    """Ensure every record has every key (None for missing)."""
    return {k: record.get(k, None) for k in ALL_KEYS}


def batch_insert(records):
    """Insert records in batches. Returns (success_count, error_count)."""
    # Normalise all records to have the same keys
    records = [normalise_keys(r) for r in records]
    success = 0
    errors = 0
    total = len(records)

    for start in range(0, total, BATCH_SIZE):
        chunk = records[start : start + BATCH_SIZE]
        try:
            resp = requests.post(
                f"{SUPABASE_URL}/rest/v1/orders?on_conflict=case_code",
                headers=HEADERS,
                data=json.dumps(chunk, default=str),
                timeout=30,
            )
            if resp.status_code in (200, 201):
                success += len(chunk)
                print(f"  Inserted rows {start + 1}–{start + len(chunk)} of {total} ✓")
            else:
                errors += len(chunk)
                print(f"  ERROR rows {start + 1}–{start + len(chunk)}: {resp.status_code} {resp.text[:200]}")
        except Exception as exc:
            errors += len(chunk)
            print(f"  EXCEPTION rows {start + 1}–{start + len(chunk)}: {exc}")

    return success, errors


def main():
    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    all_records = []

    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet '{sheet_name}' not found — skipping")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"  Sheet '{sheet_name}' is empty — skipping")
            continue

        # First row is header — skip it
        data_rows = rows[1:]
        sheet_records = []  # type: ignore
        skipped = 0

        for row in data_rows:
            record = row_to_order(row)
            if record is None:
                skipped += 1
                continue
            sheet_records.append(record)

        print(f"\nSheet '{sheet_name}': {len(sheet_records)} valid rows, {skipped} empty/skipped")
        all_records.extend(sheet_records)

    # Deduplicate by case_code — keep the last occurrence
    seen = {}
    for r in all_records:
        key = r.get("case_code") or id(r)
        seen[key] = r
    all_records = list(seen.values())

    print(f"\nTotal records to insert (after dedup): {len(all_records)}")

    if not all_records:
        print("Nothing to insert.")
        sys.exit(0)

    print("\nInserting into Supabase...")
    success, errors = batch_insert(all_records)

    print(f"\n--- Import complete ---")
    print(f"  Inserted: {success}")
    print(f"  Errors:   {errors}")
    print(f"  Total:    {len(all_records)}")


if __name__ == "__main__":
    main()
